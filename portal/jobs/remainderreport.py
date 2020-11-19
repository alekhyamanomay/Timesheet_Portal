import os
import openpyxl
from openpyxl.styles.borders import Border, Side
import init
from portal.models.remainders import Remainders
from datetime import datetime, timedelta
from portal.models import db
fromdate = datetime.now().date() + timedelta(days= -7)
todate = datetime.now().date()

print(fromdate,todate)


query = db.session.query(Remainders.UserId,Remainders.UserName, db.func.count(Remainders.UserId)).filter(Remainders.RemainderDate >= fromdate,Remainders.RemainderDate <= todate).group_by(Remainders.UserId,Remainders.UserName).all()
print(query)

datestring = str(fromdate) +"_TO_"+str(todate)
filename = f"Remainder_Report_{datestring}.xlsx"
file_path = os.path.abspath(os.path.join(init.app.config['ROOT_DIR'],"portal","jobs","Templates","RemainderReport.xlsx"))
save_path = os.path.abspath(os.path.join(init.app.config['REPORT_DIR'],"RemainderReports",filename))
print("filename",filename)
print("file_path",file_path)
print("save_path",save_path)
book = openpyxl.load_workbook(file_path)
sheet = book.active
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
row = 2
for value in query:
    sheet.cell(row=row, column=1).value =  value[0]
    sheet.cell(row=row, column=1).border = thin_border
    sheet.cell(row=row, column=2).value =  value[1]
    sheet.cell(row=row, column=2).border = thin_border
    sheet.cell(row=row, column=3).value =  value[2]
    sheet.cell(row=row, column=3).border = thin_border
    row +=1

book.save(save_path)
