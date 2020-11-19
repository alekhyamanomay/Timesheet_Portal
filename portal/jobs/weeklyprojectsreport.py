import os
from datetime import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import pyodbc
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=DESKTOP-IRHNO1M\SQLEXPRESS;'
                      'Database=TimeSheet;'
                      "UID=Manomay1;"
                      "PWD=Manomay@9;"
                      'Trusted_Connection=no;')

thin_border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
right=Side(border_style=BORDER_THIN, color='00000000'),
top=Side(border_style=BORDER_THIN, color='00000000'),
bottom=Side(border_style=BORDER_THIN, color='00000000')
)
side = Border(right=Side(border_style=None))
cur = conn.cursor()
select_query = "select Customer, Project, WeekDate, UserId, UserName, TaskName, SubTaskName, Timespent, Description from Timesheetentry group by Customer, Project, WeekDate, UserId, UserName, TaskName, SubTaskName, Timespent, Description having WeekDate between DateAdd(Day,-7,CAST(getdate() as date)) and DateAdd(Day,-1,CAST(getdate() as date))"
rows = cur.execute((select_query))
row = rows.fetchall()

td = date.today()
t_month_day_year = td.strftime("%d%m%y")

header = ('Customer', 'Project', 'Date','UserId', 'UserName', 'Task', 'SubTask', 'Time Spent', 'Description')

if row:
    prev = None
    prev1 = None
    curr = None
    curr1 = None
    list1 = []
    wb = Workbook()
    ws = wb.active
    file_name = 'Projectweekly' + '_' + t_month_day_year + '.xlsx'
    for r1 in row:
        curr = r1[0]
        curr1 = r1[1]
        if prev == None:
            print("in none")
            #file_name = 'Projectweekly' + r1[0] + '_' + t_month_day_year + '.xlsx'
            file_name = 'C:\\Users\\Manomay\\Desktop\\Reports\\ProjectReports\\' + r1[0] + '_' + t_month_day_year + '.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = "sheet1"
            std1 = wb["sheet1"]
            wb.remove(std1)
            r2 = [r1[0], r1[1], r1[2], int(r1[3]), r1[4], r1[5], r1[6], float(r1[7]), r1[8]]
            prev = r1[0]
            list1.append(r2)
        elif prev != curr:
            r2 = [r1[0], r1[1], r1[2], int(r1[3]), r1[4], r1[5], r1[6], float(r1[7]), r1[8]]
            #list1.append(r2)
            print("here")
            prev1 = None
            #print(list1)
            for each1 in list1:
                if prev1 == None:
                    
                    ws = wb.create_sheet(each1[1])
                    ws.append((header))
                    ws['A1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['A1'].font = Font(bold=True)
                    ws['B1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['B1'].font = Font(bold=True)
                    ws['C1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['C1'].font = Font(bold=True)
                    ws['D1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['D1'].font = Font(bold=True)
                    ws['E1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['E1'].font = Font(bold=True)
                    ws['F1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['F1'].font = Font(bold=True)
                    ws['G1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['G1'].font = Font(bold=True)
                    ws['H1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['H1'].font = Font(bold=True)
                    ws['I1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['I1'].font = Font(bold=True)
                    # print("1", each1)
                    ws.append(each1)
                    prev1 = each1[1]
                elif each1[1] == prev1:
                    ws.append(each1)
                    prev1 = each1[1]
                else:
                    prev1 = each1[1]
                    # ws.append(each1)
                    # for each column adjust the column width to the length of the maximum value in a cell
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:  # Necessary to avoid error on empty cells
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')
                        adjusted_width = (max_length + 2) * 1.1
                        ws.column_dimensions[column].width = adjusted_width
                    ws.sheet_view.showGridLines = False

                    ws = wb.create_sheet(each1[1])
                    ws.append(header)
                    ws.append(each1)

                    ws['A1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['A1'].font = Font(bold=True)
                    ws['B1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['B1'].font = Font(bold=True)
                    ws['C1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['C1'].font = Font(bold=True)
                    ws['D1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['D1'].font = Font(bold=True)
                    ws['E1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['E1'].font = Font(bold=True)
                    ws['F1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['F1'].font = Font(bold=True)
                    ws['G1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['G1'].font = Font(bold=True)
                    ws['H1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['H1'].font = Font(bold=True)
                    ws['I1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
                    ws['I1'].font = Font(bold=True)

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = adjusted_width
            ws.sheet_view.showGridLines = False

            #wb.remove(std)
            wb.save(file_name)
            wb.close()

            #for each in list1:
            #    ws.append((each))
            list1=[]
            list1.append(r2)
            wb.save(file_name)
            wb.close()
            #file_name = 'Projectweekly' + r1[0] + '_' + t_month_day_year + '.xlsx'
            file_name = 'C:\\Users\\Manomay\\Desktop\\Reports\\ProjectReports\\' + r1[0] + '_'+t_month_day_year+'.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = "sheet1"
            std = wb['sheet1']
            wb.remove(std)
            prev = r1[0]


        elif prev == curr:
            print("in equal")
            r2 = [r1[0], r1[1], r1[2], int(r1[3]), r1[4], r1[5], r1[6], float(r1[7]), r1[8]]
            prev = r1[0]
            list1.append((r2))

    for each1 in list1:
        if prev1 == None:
            ws = wb.create_sheet(each1[1])
            ws.append((header))
            ws['A1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['A1'].font = Font(bold=True)
            ws['B1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['B1'].font = Font(bold=True)
            ws['C1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['C1'].font = Font(bold=True)
            ws['D1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['D1'].font = Font(bold=True)
            ws['E1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['E1'].font = Font(bold=True)
            ws['F1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['F1'].font = Font(bold=True)
            ws['G1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['G1'].font = Font(bold=True)
            ws['H1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['H1'].font = Font(bold=True)
            ws['I1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['I1'].font = Font(bold=True)
            #print("1", each1)
            ws.append(each1)
            prev1 = each1[1]
        elif each1[1] == prev1:
            ws.append(each1)
            prev1 = each1[1]
        else:
            prev1 = each1[1]
            #ws.append(each1)
            # for each column adjust the column width to the length of the maximum value in a cell
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = adjusted_width
            ws.sheet_view.showGridLines = False


            ws = wb.create_sheet(each1[1])
            ws.append(header)
            ws.append(each1)

            ws['A1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['A1'].font = Font(bold=True)
            ws['B1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['B1'].font = Font(bold=True)
            ws['C1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['C1'].font = Font(bold=True)
            ws['D1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['D1'].font = Font(bold=True)
            ws['E1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['E1'].font = Font(bold=True)
            ws['F1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['F1'].font = Font(bold=True)
            ws['G1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['G1'].font = Font(bold=True)
            ws['H1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['H1'].font = Font(bold=True)
            ws['I1'].fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type='solid')
            ws['I1'].font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width
    ws.sheet_view.showGridLines = False


    #wb.remove(std)

    wb.save(file_name)
    wb.close()
    #wb.remove(std)
    #print(list1)


if os.path.exists('Projectweekly' + '_' + t_month_day_year + '.xlsx'):
    os.remove('Projectweekly' + '_' + t_month_day_year + '.xlsx')
else:
    print("file doesn't exist")


cur.close()
conn.close()
