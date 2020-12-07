import os
import threading

import openpyxl
from datetime import datetime
from flask import send_file
from flask_restx import Resource, inputs, reqparse
from openpyxl.styles.borders import Border, Side
from portal.helpers import delete_excel, token_decode, token_verify_or_raise
from portal.models.tasks import Tasks
from portal.models.timesheetentry import TimesheetEntry
from werkzeug.exceptions import UnprocessableEntity

from ... import APP
from . import ns

parser = reqparse.RequestParser()
# parser.add_argument('Authorization', type=str,location='headers', required=True)
parser.add_argument('userid', type=int,location='json', required=False)
parser.add_argument('project', type=str,location='json', required=False)
parser.add_argument('customer', type=str,location='json', required=False)
parser.add_argument('fromdate', type=inputs.date_from_iso8601,location='json', required=True)
parser.add_argument('todate', type=inputs.date_from_iso8601,location='json', required=True)


@ns.route('/project')
class Get_values(Resource):
    @ns.doc(description='Generate Project Report',
            responses={200: 'OK', 400: 'Bad Request', 401: 'Unauthorized', 500: 'Internal Server Error'})
    @ns.expect(parser, validate=True)
    def get(self):
        args = parser.parse_args(strict=True)
        try:
            # y = token_decode(args['Authorization'])
            # if isinstance(y,tuple):
            #     return {'error':"Unathorized token"}, 401
            # token_verify_or_raise(args['Authorization'])
            print(args)
            query = TimesheetEntry.query
            if args['project'] and args['customer'] is not None:
                query = query.filter(TimesheetEntry.Project == args['project'], TimesheetEntry.Customer==args['customer'])
            if args['userid'] is not None:
                query = query.filter(TimesheetEntry.UserId == args['userid'])
            query = query.filter(TimesheetEntry.WeekDate >= args['fromdate'], TimesheetEntry.WeekDate <= args['todate']).all()
            datestring = str(args['fromdate'].strftime("%d-%m-%Y")) +"_TO_"+str(args['todate'].strftime("%d-%m-%Y"))
            filename = f"Project_Report_{datestring}_{str(datetime.now().timestamp()).replace('.','')}.xlsx"
            file_path = os.path.abspath(os.path.join(APP.config['ROOT_DIR'],"Templates","sheets","ProjectReportTemplate.xlsx"))
            save_path = os.path.abspath(os.path.join(APP.config['ROOT_DIR'],"Templates","Temp",filename))
            book = openpyxl.load_workbook(file_path)
            sheet = book.active
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
            row = 3
            for value in query:
                sheet.cell(row=row, column=1).value =  value.Customer
                sheet.cell(row=row, column=2).value =  value.Project
                sheet.cell(row=row, column=3).value =  value.WeekDate
                sheet.cell(row=row, column=4).value =  value.UserId
                sheet.cell(row=row, column=5).value =  value.UserName
                sheet.cell(row=row, column=6).value =  value.TaskName
                sheet.cell(row=row, column=7).value =  value.SubTaskName
                sheet.cell(row=row, column=8).value =  value.Timespent
                sheet.cell(row=row, column=9).value =  value.Description
                row +=1
            # adjusting the column widths
            for col in sheet.columns:
                    max_length = 0
                    column = col[2].column_letter
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                        cell.border = thin_border
                    adjusted_width = (max_length + 2) * 1.1
                    sheet.column_dimensions[column].width = adjusted_width
            book.save(save_path)
            threading.Thread(target=delete_excel, args=(save_path,)).start()
            return send_file(save_path, as_attachment=filename)
        except Exception as e:
            APP.logger.exception("some error occurred")
            return UnprocessableEntity('Some error occurred')
