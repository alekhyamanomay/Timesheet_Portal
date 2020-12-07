import os
import threading

import openpyxl
from datetime import datetime
from flask import send_file
from flask_restx import Resource, inputs, reqparse
from openpyxl.styles.borders import Border, Side
from portal.helpers import delete_excel
from portal.models.remainders import Remainders
from portal.models import db
from werkzeug.exceptions import UnprocessableEntity

from ... import APP
from . import ns

parser = reqparse.RequestParser()
# parser.add_argument('Authorization', type=str,location='headers', required=True)
parser.add_argument('userid', type=int,location='json', required=False)
parser.add_argument('fromdate', type=inputs.date_from_iso8601,location='json', required=True)
parser.add_argument('todate', type=inputs.date_from_iso8601,location='json', required=True)


@ns.route('/remainder')
class Get_values(Resource):
    @ns.doc(description='Generate Project Report',
            responses={200: 'OK', 400: 'Bad Request', 401: 'Unauthorized', 500: 'Internal Server Error'})
    @ns.expect(parser, validate=True)
    # @ns.marshal_with(response_model)
    def get(self):
        args = parser.parse_args(strict=True)
        try:
            # y = token_decode(args['Authorization'])
            # if isinstance(y,tuple):
            #     return {'error':"Unathorized token"}, 401
            # token_verify_or_raise(args['Authorization'])
            print(args)
            fromdate = args['fromdate']
            todate = args['todate']

            print(fromdate,todate)

            query = db.session.query(Remainders.UserId,Remainders.UserName, db.func.count(Remainders.UserId)).filter(Remainders.RemainderDate >= fromdate,Remainders.RemainderDate <= todate)
            if args['userid'] is not None:
                query = query.filter(Remainders.UserId == 1058)
            query = query.group_by(Remainders.UserId,Remainders.UserName).all()
            print(query)
            datestring = str(fromdate.strftime("%d-%m-%Y")) +"_TO_"+str(todate.strftime("%d-%m-%Y"))
            filename = f"Remainder_Report_{datestring}_{str(datetime.now().timestamp()).replace('.','')}.xlsx"
            file_path = os.path.abspath(os.path.join(APP.config['ROOT_DIR'],"Templates","sheets","RemainderReport.xlsx"))
            save_path = os.path.abspath(os.path.join(APP.config['ROOT_DIR'],"Templates","Temp",filename))
            print("filename",filename)
            print("file_path",file_path)
            print("save_path",save_path)
            book = openpyxl.load_workbook(file_path)
            sheet = book.active
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
            row = 2
            for value in query:
                sheet.cell(row=row, column=1).value =  value[0]
                sheet.cell(row=row, column=1).border = thin_border
                sheet.cell(row=row, column=2).value =  value[1]
                sheet.cell(row=row, column=2).border = thin_border
                sheet.cell(row=row, column=3).value =  value[2]
                sheet.cell(row=row, column=3).border = thin_border
                row +=1

            book.save(save_path)
            threading.Thread(target=delete_excel, args=(save_path,)).start()
            return send_file(save_path, as_attachment=filename)
        except:
            APP.logger.exception("some error occurred")
            return UnprocessableEntity('Some error occurred')